# PRG1:ライブラリ設定
import sys
import os
import time
import shutil
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import io
from PIL import Image
import openpyxl as px
from openpyxl.styles import Alignment
from selenium.webdriver.chrome import service as fs

import const.status as status
import const.category as category
 
# メインプログラム
def main():
    print('--- START ---')
    # PRG2:クロール設定
    
    CUR_DIR = os.getcwd()
    args = sys.argv
    KEYWORD = args[1]
    for i in range(2,len(args)):
        KEYWORD += '+' + args[i]

    BASE_URL = 'https://jp.mercari.com'
    URL_INI = create_url(BASE_URL, KEYWORD, status.const.SOLD_OUT, category.const.POKEKA)
    url = URL_INI
    
    page_num = 1

    CHROMEDRIVER = '/opt/chrome/chromedriver'
 
    # 検索結果格納リスト・画像保存フォルダ作成
    result = []
    if os.path.isdir('./img') == False:
        os.mkdir('./img')
    
    print('scraping started')
    print('scraping url: ' + url)
    # PRG3:スクレイピング実行
    while True:
        try:
            # ChromeでURLに接続
            options = Options()
            options.add_argument('--headless')  
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')

            chrome_service = fs.Service(executable_path=CHROMEDRIVER) 
            browser = webdriver.Chrome(service=chrome_service, options=options)
            browser.get(url)
            time.sleep(5)
            html = browser.page_source.encode('utf-8')

            # Beautifulsoupで要素取得
            soup = BeautifulSoup(html, "html.parser")
            items_list = soup.find_all("li", attrs={'data-testid':'item-cell'})
 
            # サムネイル画像取得
            for i, item in enumerate(items_list):
                # ID
                id = item.find('a').get('href').split('/')[-1]
                # 商品名
                text = item.text
                # 価格
                price = int(item.find("mer-item-thumbnail").get('price'))
                # 画像URL
                img_src = item.find("mer-item-thumbnail").get('src')
                # 商品詳細URL
                detail_url =  BASE_URL + item.find('a').get('href')
                
                # 画像保存
                img_fname = save_img(img_src)
 
                # 取得結果をリストに保存
                result.append([
                    text,
                    price,
                    os.path.join(CUR_DIR, 'img', img_fname),
                    detail_url
                ])
            # 次ページボタン処理
            next_button = soup.find('mer-button',attrs={'data-testid':"pagination-next-button"})
            if next_button:
                page_num += 1
                param = '&page_token=v1%3A' + str(page_num)
                next_url = URL_INI + param
                url = next_url
                next_url = ''
            else:
                break
 
        # エラー発生時の例外処理
        except Exception as e:
            message = "[エラー]"+"type:{0}".format(type(e))+"\n"+"args:{0}".format(e.args)
            print(message)
 
        # Chrome終了処理
        finally:
            browser.close()
            browser.quit()
    
    print('excel creating...')
    create_excel(KEYWORD, result)
 
    # 保存した画像の削除
    print('temp image folder deleting...')
    shutil.rmtree('./img')
 
    print('--- END ---')
 
# PRG5:アスペクト比固定して画像リサイズ
def scale_to_width(img, width):  
    height = round(img.height * width / img.width)
    return img.resize((width, height))

# URL生成
def create_url(base_url,keyword, status, category):
    # 検索用URL生成
    url = base_url + '/search/?keyword=' + keyword + '&status=' + status + '&category_id=' + category
    return url

# Excelファイル生成
def create_excel(keyword, result):
    # PRG4:Excelに結果出力
    wb = px.Workbook()
    ws = wb.active
 
    # 書式設定
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    my_alignment=Alignment(vertical='top', wrap_text=True)
 
    # Excelヘッダー出力
    headers = ['アイテム名','価格','サムネイル','URL']
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1, value=header)
 
    # 取得結果書き込み
    for y, row in enumerate(result):
        ws.row_dimensions[y+2].height = 160
        for x, cell in enumerate(row):
            if x == 2:
                img = px.drawing.image.Image(cell)
                img.anchor = ws.cell(row= y+2, column= x+1).coordinate
                ws.add_image(img)
            elif x == 3:
                ws.cell(row= y+2, column= x+1).hyperlink = cell
                ws.cell(row= y+2, column= x+1).alignment = my_alignment
            else:
                ws.cell(row= y+2, column= x+1).value = cell
                ws.cell(row= y+2, column= x+1).alignment = my_alignment
            
    # Excelファイル保存
    xlname = './mercari_' + keyword + '.xlsx'
    wb.save(xlname)

# 画像保存
def save_img(img_src):
    response = requests.get(img_src)
    img_fname = img_src.split('?')[0].split('/')[-1]
    img_bin = io.BytesIO(response.content)
    pil_img = Image.open(img_bin)
    img_resize = scale_to_width(pil_img, 200)
    img_resize.save('./img/' + img_fname)
    return img_fname

# スクリプトとして実行された場合の処理
if __name__ == '__main__':
    main()